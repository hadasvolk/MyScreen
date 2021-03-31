# -*- coding: utf-8 -*-
import sys
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook

try:
    import cfg
    import libs.tools as tools
except:
    sys.path.append(r'c:\Gamidor\MyScreen\Script\src')
    sys.path.append(r'c:\Gamidor\MyScreen\Script\src\libs')
    import cfg
    import tools


csv_cols = ['Sample', 'Test Code', 'Test Name', 'Disease', 'Gene', 'Mutation',
            'MOH', 'Ethnicity', 'Classification', 'Clalit Disease Makat',
            'Clalit Mutation Makat', 'Remark', 'GQX', 'Alt Variant Freq',
            'Read Depth', 'Alt Read Depth', 'Allelic Depths', 'Correlation',
            'N comp', 'Custom First', 'Custom Last', 'BF', 'Reads Expected',
            'Reads Observed', 'Reads Ratio', 'Analyzed Gender', 'Reported Gender',
            'Sample Source', 'Mother Ethnicity', 'Father Ethnicity',
            'Partner Sample','AGID', 'IGV Link (open IGV before)', 'Result 2']

# DMD mutation makats
DMD = cfg.DMD


def split_rows(df):

    s = df["AGID"].str.split('/', expand=True).stack()
    i = s.index.get_level_values(0)
    df2 = df.loc[i].copy()
    df2["AGID"] = s.values

    return df2


def get_panels_agids():

    panels_agids = {}
    for panel in cfg.Panels_names:
        try:
            panel_set = tools.decompress_pickle(
                "{}/{}.AGID.reported.pbz2".format(cfg.path_to_panel, panel))
            panels_agids[panel] = pd.DataFrame(list(panel_set), columns=['AGID'])
            panels_agids[panel].set_index('AGID', inplace=True)
        except:
            pass

    return panels_agids


def get_annos():

    geno_anno = pd.read_pickle(cfg.anno_pkl)
    geno_anno_wt = geno_anno[geno_anno['Classification'] == 'WT'].copy()
    geno_anno_wt['Mutation'] = geno_anno_wt[['annotation2', 'gdna']].agg('\n'.join, axis=1)
    ganno = geno_anno_wt[['AGID', 'Mutation']]
    ganno.set_index('AGID', inplace=True)

    cnv_anno = pd.read_pickle(cfg.annotCNV)
    cnv_anno['Disease'] = cnv_anno['Annotation1'].apply(lambda x: x.split(':')[1])
    cnv_anno['Mutation'] = cnv_anno[['Annotation2', 'gdna']].agg('\n'.join, axis=1)
    cnv_anno.rename(columns={'Annotation3':'MOH', 'Annotation4':'Ethnicity'}, inplace=True)
    cnv_anno[['Correlation', 'N comp', 'Custom First', 'Custom Last', 'BF', 'Reads Expected', 'Reads Observed', 'Reads Ratio']] = np.nan
    cnv_anno.drop(columns=['Chr', 'Start', 'Stop', 'gdna', 'Custom.first', 'Custom.last', 'Annotation1', 'Annotation2', 'del_length'], inplace=True)
    cnv_anno['Classification'], cnv_anno['Genotype'] = 'WT', 'hom'
    cnv_anno.set_index('AGID', inplace=True)

    return ganno, cnv_anno


def explode(df, lst_cols, fill_value='', preserve_index=False):
    # make sure `lst_cols` is list-alike
    if (lst_cols is not None
        and len(lst_cols) > 0
        and not isinstance(lst_cols, (list, tuple, np.ndarray, pd.Series))):
        lst_cols = [lst_cols]
    # all columns except `lst_cols`
    idx_cols = df.columns.difference(lst_cols)
    # calculate lengths of lists
    lens = df[lst_cols[0]].str.len()
    # preserve original index values    
    idx = np.repeat(df.index.values, lens)
    # create "exploded" DF
    res = (pd.DataFrame({
                col:np.repeat(df[col].values, lens)
                for col in idx_cols},
                index=idx)
             .assign(**{col:np.concatenate(df.loc[lens>0, col].values)
                            for col in lst_cols}))
    # append those rows that have empty lists
    if (lens == 0).any():
        # at least one list in cells is empty
        res = (res.append(df.loc[lens==0, idx_cols], sort=False)
                  .fillna(fill_value))
    # revert the original index order
    res = res.sort_index()
    # reset index if requested
    if not preserve_index:        
        res = res.reset_index(drop=True)
    return res


def create_summary_csv(**data):

    All_split = split_rows(tools.decompress_pickle(data["All"]))
    All_split = All_split[All_split.AGID != 'AGID Not Found']
    All_split[['Sample', 'S']] = All_split.Sample.str.split('_S',expand=True)
    All_split['Disease'] = All_split['Custom Annotation'].apply(lambda x : x.split(':')[1])
    All_split.rename(columns={'Custom Annotation 3':'MOH', 'Custom Annotation 4':'Ethnicity'}, inplace=True)
    All_split.drop(columns=['Coordinate', 'Variant', 'Chr', 'gdna', 'Custom Annotation', 'Custom Annotation 2', 'Type', 'SoftClipped Reads'], inplace=True)
    All_split.set_index('AGID', inplace=True)

    wb = load_workbook(filename = data["summary"])
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    summary = pd.DataFrame(ws.values)
    summary.columns = summary.iloc[4]
    summary = summary.iloc[5:]
    summary['AGID'] = summary['AGID'].apply(lambda x: x.split('/'))
    summary = explode(summary, ['AGID'])
    summary.reset_index(inplace=True)
    summary = summary[summary.AGID != 'AGID Not Found']
    summary.set_index('AGID', inplace = True)

    summary_csv = pd.DataFrame(columns=summary.columns)

    ganno, cnv_anno = get_annos()

    annotated = All_split.join(ganno, how='left')

    panels_agids = get_panels_agids()

    same_cols = set(summary.columns) - set(annotated.columns) - set(cnv_anno.columns)
    same_cols.remove('IGV Link (open IGV before)')

    for sample in set(annotated.Sample.tolist()):
        sample =str(sample)
        cur = annotated[annotated.Sample == sample]
        cur = pd.concat([cur, cnv_anno])

        cur['Sample'] = sample
        cur['S'] = cur.iloc[0]['S']

        in_positive = summary[summary['Sample'].astype('str') == sample]
        in_positive = in_positive[~in_positive.index.duplicated(keep='first')]
        panel = in_positive.iloc[0]['Test Name']
        cur_panel = panels_agids[panel].join(cur, how='left')
        not_in_panel = [x for x in in_positive.index.tolist() if x not in cur_panel.index.tolist()] 
        try:
            not_in_panel.remove('AG0000')
        except ValueError:
            pass  # do nothing!
        in_positive.drop(not_in_panel, inplace=True)
        pos = in_positive.index.tolist()
        
        if pos != ['AG0000']:
            pos = [x.split('_')[0] for x in pos]
            cur_panel.drop(pos, inplace=True)
            cur_panel = pd.concat([in_positive, cur_panel])

        for col in same_cols:
            cur_panel[col] = in_positive.iloc[0][col]

        if in_positive.iloc[0]['Classification'] == 'SAMPLE FAILED':
            cur_panel['Classification'] = 'SAMPLE FAILED'

        cur_panel.reset_index(inplace=True)
        cur_panel['AGID'] = cur_panel['AGID'].apply(lambda x: x.split('_')[0])
        dmd = cur_panel[cur_panel.Gene == 'DMD']
        for ag in dmd.AGID.unique():
            dmd_ag = dmd[dmd.AGID == ag]
            for i,(index,row) in enumerate(dmd_ag.iterrows(), 1):
                cur_panel.loc[index, 'AGID'] = f'{row.AGID}_{i}'
                cur_panel.loc[index, 'Clalit Mutation Makat'] = DMD[f'{row.AGID}_{i}']
                cur_panel.loc[index, 'Result 2'] = cur_panel.loc[index, 'Mutation']
        cur_panel.set_index('AGID', inplace=True)

        summary_csv = summary_csv.append(cur_panel)

    summary_csv.rename(columns={'Genotype':'Remark'}, inplace=True)
    summary_csv.index.rename('AGID', inplace=True)
    summary_csv.reset_index(inplace=True)
    summary_csv.fillna('---', inplace=True)
    summary_csv = summary_csv.replace('\n',' ', regex=True)

    def split_join(x):
        try:
            return ' '.join(x.split())
        except:
            return x

    string_cols = summary_csv.dtypes[summary_csv.dtypes == 'object'].index.tolist()
    for col in string_cols:
        summary_csv[col] = summary_csv[col].apply(lambda x: split_join(x))

    header = ",Data Analysis Version,,Run Name,,Analysis Date,,,,,,,,,,,,,,,,,,,,,,,,,,,\n \
        ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,\n \
        ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,\n \
        ,{},,{},,{},,,,,,,,,,,,,,,,,,,,,,,,,,, \n \
        ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,, \n \
        ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,, \n".format(data["v"], data["r"], data["d"])

    csv_file = "{}/sample_summary-{}.csv".format(data["out"], data["d"])
    with open(csv_file, 'w', newline='', encoding='utf-8') as fp:
        fp.write(header)
        summary_csv[csv_cols].to_csv(fp, index=False)

    psv_file = "{}/sample_summary-{}.psv".format(data["out"], data["d"])
    with open(psv_file, 'w', newline='', encoding='utf-8') as fp:
        fp.write(header.replace(',', '|'))
        summary_csv[csv_cols].to_csv(fp, index=False, sep="|")

    return csv_file, psv_file



def csv_dumper(sample_summary_excel):
    """ Deprected legacy CSV creator"""

    anno_geno = pd.read_pickle(cfg.anno_pkl)
    anno_cnv = pd.read_pickle(cfg.annotCNV)

    raw = pd.read_excel(sample_summary_excel, sheet_name='Sheet1', header=4)
    UniqueSamples = raw.Sample.unique()
    dfDict = {elem : pd.DataFrame for elem in UniqueSamples}
    for key in dfDict.keys():
        dfDict[key] = raw[:][raw.Sample == key]
    for k,v in dfDict.items():
        dfDict[k].reset_index(inplace=True)
        cases = dfDict[k].AGID.unique()
        print(cases)
        panel = dfDict[k].loc[0, 'Test Name']
        try:
            ags = tools.decompress_pickle(f'{cfg.path_to_panel}{panel}.AGID.pbz2')
        except:
            ags_df = pd.read_csv(cfg.ver2_agids)
            ags = ags_df.AGID.unique()
        ags_filter = [x for x in ags if "/" not in x]
        to_keep = []
        for agid in ags_filter:
            sub = anno_geno[anno_geno.AGID == agid].dropna(how='all')
            classes = sub.Classification.unique()
            if any('CARRIER' in x for x in classes) or not len(classes):
                to_keep.append(agid)
            else:
                if agid in cases:
                    print(sub.Classification.unique(), agid)
        agids = pd.DataFrame(to_keep, columns=['AGID'])

        # agids = [agid for line in agids for agid in line.split('/')]
        # print(agids)
        # print(k, panel)
        # print(cases)
        full = dfDict[k].merge(agids, on='AGID', how='outer').reset_index()
        # print(full)
        full.to_csv('test.csv')
        print()
        # break


if __name__ == "__main__":

    all_path = r'c:\Users\hadas\AGcloud\AGshared\Gamidor\Capture_Panel\HospitalRuns\Zer\210324_MN00742_0120_A000H3F7NM\MyScreen_Analysis_v2.2_RESULTS\Info\Genotyping\Logs\All.filtered.pbz2'
    summary_path = r'c:\Users\hadas\AGcloud\AGshared\Gamidor\Capture_Panel\HospitalRuns\Zer\210324_MN00742_0120_A000H3F7NM\MyScreen_Analysis_v2.2_RESULTS\sample_summary-26-03-2021.xlsx'

    ver = "MyScreen_Analysis_v2.2"
    run = "210324_MN00742_0120_A000H3F7NM"
    date = "30-Nov-20"
    o = os.getcwd()

    create_summary_csv(All = all_path, summary = summary_path, v = ver, r = run,
        d = date, out = o)

    # csv_dumper(r'c:\Users\hadas\AGcloud\AGshared\Gamidor\Capture_Panel\HospitalRuns\Zer\201122_MN00742_0106_A000H37LVW\MyScreen_Analysis_v2.1_RESULTS\sample_summary-26-03-2021.xlsx')
