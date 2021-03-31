import sys
import pandas as pd
import csv
import openpyxl
import openpyxl.styles as opstyl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, \
    colors  # for IGV link (only colors)
import matplotlib
import matplotlib.cm as cm
import numpy as np
import random
import string
from openpyxl import load_workbook
from itertools import islice
import glob
import shutil
import datetime
import re  # for igv link
import os  # for igv link
import subprocess  # for igv link
import pickle  # for igv link

try:
    import cfg
    import libs.tools as tools
except:
    sys.path.append(r'c:\Gamidor\MyScreen\Script\src')
    sys.path.append(r'c:\Gamidor\MyScreen\Script\src\libs')
    import cfg
    import tools

time = datetime.datetime.now()
date = datetime.datetime.now().strftime("%d-%m-%Y")

black = 'FF000000'
red = 'FFFF0028'
yellow = 'ffff00'
orange = 'ffa500'
cyan = '00FFFF'
grey = 'E0E0E0'
magneta = 'C5D9F1'
purple_softclipped = 'E4C7F0'
green = 'D2F9C7'

thin = Side(border_style="thin", color="000000")
thick = Side(border_style="thick", color="000000")
double = Side(border_style="double", color="000000")

georgian_druze = ['CARRIER-Georgian', 'CARRIER-Georgian - With soft-clipped reads', 'CARRIER-Druze']
non_reported_wt = 'WT - Low GQX - NON_REPORTED variant in the same loc'
problems = ['WT', 'WT-Problem', 'Problem', 'NO_CALL', 'NO_CALL-Problem', 'CARRIER-Problem',
            'WT - With soft-clipped reads', 'WT-Problem - With soft-clipped reads',
            'Problem - With soft-clipped reads', 'NO_CALL - With soft-clipped reads',
            'CARRIER-Problem - With soft-clipped reads', 'CARRIER-Georgian-Problem',
            'CARRIER-Georgian-Problem - With soft-clipped reads', 'CARRIER-Druze-Problem']
problems_cnv = ['CNV-Problem', 'CNV-Problem Big Del Boundaries Different as Reported']
warnings = ['Big Del Boundaries Different as Reported',
            'CNV Big Del Boundaries Different as Reported']


def fill(cell, color, type='solid'):
    cell.fill = opstyl.PatternFill(start_color=color,
                                   end_color=color,
                                   fill_type=type)


def adjust_df(df, paths):
    # df.sort_values(by=['S'], inplace=True)
    df.reset_index(inplace=True)
    df['S'] = pd.to_numeric(df['S'])
    df.sort_values(by=['S'], inplace=True)
    run = df.iloc[1]['Date of Run']
    df = df.drop(columns=['Date of Run'])
    # df.drop(columns=['S'], inplace=True)

    cols = list(df.columns.values)
    change = ['Sex', 'Sample Source', 'Mother Ethnicity', 'Father Ethnicity',
              'Partner Sample', 'AGID']
    for col in change:
        cols.pop(cols.index(col))
    df = df[cols + change]
    df.drop(columns=['Genotype'], inplace=True)
    df.rename(columns={"Sex": "Analyzed Gender", "Gender": "Reported Gender",
                       "N.comp": "N comp", "Custom.first": "Custom First",
                       "Custom.last": "Custom Last", "Reads.expected": "Reads Expected",
                       "Reads.observed": "Reads Observed", "Reads.ratio": "Reads Ratio",
                       "Classification" : "Genotype"}, inplace=True)
    df["Classification"] = np.nan
    df["Test Code"] = np.nan
    df["Test Name"] = np.nan

    try:
        failed = tools.decompress_pickle("{}/failed_samples.pbz2".format(paths["DIR_TREE"][6]))
        samples = [f.split('_')[0] for f in failed]
    except Exception as e:
        raise e

    D = {name : code for name, code in cfg.Panels}
    for index, row in df.iterrows():
        s = 'S{}'.format(row.S)
        panel = paths['SAMPLE_DICT'][s][1]
        df.loc[index, "Test Code"] = D.get(panel)
        df.loc[index, "Test Name"] = panel
        if row.Disease == "No mutations identified (WT).":
            df.loc[index, "AGID"] = "AG0000"
            df.loc[index, "Classification"] = "No mutations identified (WT)."
        if row.Sample in samples:
            if row.Disease == "No mutations identified (WT).":
                df.loc[index, "Disease"] = "SAMPLE FAILED"
                # df.loc[index, "AGID"] = "999999"
            df.loc[index, "Classification"] = "SAMPLE FAILED"
            df.loc[index, "Genotype"] += " and SAMPLE FAILED in CNV analysis"

    df = df[['Sample', 'S', 'Test Code', 'Test Name', 'Disease', 'Gene', 'Mutation',
            'MOH', 'Ethnicity', 'Classification', 'Clalit Disease Makat',
            'Clalit Mutation Makat', 'Genotype', 'GQX', 'Alt Variant Freq',
            'Read Depth', 'Alt Read Depth', 'Allelic Depths', 'Correlation',
            'N comp', 'Custom First', 'Custom Last', 'BF', 'Reads Expected',
            'Reads Observed', 'Reads Ratio', 'Analyzed Gender', 'Reported Gender',
            'Sample Source', 'Mother Ethnicity', 'Father Ethnicity', 'Partner Sample',
            'AGID', 'IGV Link (open IGV before)']]

    for s in df.S.unique():
        tmp = df[(df.S == s) & (df.Gene == 'DMD')].dropna(how='all')
        if not tmp.empty:
            for ag in tmp.AGID.unique():
                tmp_ag = tmp[tmp.AGID == ag].dropna(how='all')
                for i,(index,row) in enumerate(tmp_ag.iterrows(), 1):
                    df.loc[index, 'AGID'] = f'{row.AGID}_{i}'
                    # if i == 0:
                    #     continue
                    # else:
                    #     df.loc[index, 'AGID'] = f'{row.AGID}_{i}'

    return df, run


def header(cell, to_bold):
    if to_bold:
        cell.font = opstyl.Font(bold=True)
    cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    fill(cell, magneta)


def excel_formatter(df, paths, Analysis_Version):
    write_path = paths["DIR_TREE"][0]
    # tools.compressed_pickle("{}/summary".format(paths["DIR_TREE"][-1]), df)
    # print(paths)
    # wb = load_workbook(filename=write_path)
    # ws = wb.active
    # data = ws.values
    # cols = next(data)[0:]
    # data = list(data)
    # idx = [r[0] for r in data]
    # data = (islice(r, 0, None) for r in data)
    # df = pd.DataFrame(data, index=idx, columns=cols)
    df, run = adjust_df(df, paths)

    new_order = list(df.columns)
    if 'IGV Link (open IGV before)' in new_order:  # move 'IGV Link' column to the end of the table
        new_order.remove('IGV Link (open IGV before)')
        new_order.append('IGV Link (open IGV before)')
    df = df.reindex(columns=new_order)

    out_file = "{}/sample_summary-{}.xlsx".format(write_path, date)
    with pd.ExcelWriter(out_file) as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)

    wb = openpyxl.load_workbook(out_file)
    ws = wb["Sheet1"]

    float_rows = ['M', 'N', 'O', 'V', 'Y', 'R']
    for row in range(3, df.shape[0] + 2):
        for char in float_rows:
            try:
                val = round(float(ws["{}{}".format(char, row)].value), 4)
                ws["{}{}".format(char, row)].value = val
                ws["{}{}".format(char, row)].alignment = Alignment(horizontal='left')
            except:
                pass

    int_rows = ['J', 'K', 'W', 'X']
    for row in range(3, df.shape[0] + 2):
        for char in int_rows:
            try:
                val = int(ws["{}{}".format(char, row)].value)
                ws["{}{}".format(char, row)].value = val
                ws["{}{}".format(char, row)].alignment = Alignment(horizontal='left')
            except:
                pass

    ws.insert_rows(1, amount=4)
    ws.merge_cells('B2:D2')
    ws.merge_cells('B3:D3')
    ws.merge_cells('F2:G2')
    ws.merge_cells('F3:G3')

    ws['B2'] = "Data Analysis Version"
    ws['E2'] = "Run Name"
    ws['F2'] = "Analysis Date"
    ws['B3'] = Analysis_Version
    ws['E3'] = "{}".format(run)
    ws['F3'] = time.strftime("%d-%B-%Y")

    header_cells = ['G2', 'B2', 'C2', 'D2', 'E2', 'F2']
    info_cells = ['G3', 'B3', 'C3', 'D3', 'F3', 'E3']
    for cell in header_cells:
        header(ws[cell], True)
    for cell in info_cells:
        header(ws[cell], False)
    try:
        img = openpyxl.drawing.image.Image(cfg.MyScreen_Summary)
        img.anchor = 'I1'
        ws.add_image(img)
    except:
        ws['A2'] = "MyScreen"

    ws.freeze_panes = 'B6'
    letters = string.ascii_uppercase[7:29]
    for i in letters:
        ws.column_dimensions[i].width = 6
    ws.column_dimensions['B'].width = 3
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['R'].width = 8
    ws.column_dimensions['S'].width = 10
    ws.column_dimensions['AA'].width = 8
    ws.column_dimensions['AG'].width = 10
    ws.column_dimensions['AH'].width = 25

    samples = []
    for col in ws['A']:
        samples.append(col.value)
    samples.pop(0)
    samples.pop(0)
    samples = list(dict.fromkeys(samples))[::2]

    for cells in ws.iter_rows(min_row=6, min_col=1, max_col=34):
        sample_cell = cells[0]
        clas = cells[9]
        geno = cells[12]
        gqx = cells[13]
        altVar = cells[14]
        mut = cells[15]

        bf = cells[22]
        Reads_Observed = cells[25]

        ints = [cells[10], cells[11], cells[19], cells[20], cells[21], cells[22],
                cells[23], cells[24]]

        Analyzed_Gender = cells[26]
        Reported_Gender = cells[27]

        igv = cells[33]

        mut.alignment = Alignment(wrap_text=True)
        for i in range(2, 8):
            cells[i].alignment = Alignment(wrap_text=True)

        sample_cell.font = opstyl.Font(bold=True)
        sample_cell.border = Border(top=double, left=thin, right=thick, bottom=double)

        for each_cell in cells[1:]:
            try:
                each_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            except Exception as e:
                print(e)
        if sample_cell.value in samples:
            for each_cell in cells:
                fill(each_cell, grey)

        for each_cell in ints:
            try:
                each_cell.value = each_cell.value.split('.')[0]
            except:
                pass

        if igv.row != 5:  # don't color the header
            igv.font = Font(u='single', color=colors.BLUE)  # add blue color to the link
        try:
            if ' - With soft-clipped reads' in geno.value:
                fill(geno, purple_softclipped)
        except:
            geno.value = "unknown variant : please refer to VS2.vcf"
            cells[4].value = "unknown variant : please refer to VS2.vcf"
            cells[6].value = "unknown variant : please refer to VS2.vcf"

        if Reported_Gender.value != Analyzed_Gender.value and Reported_Gender.value != None:
            fill(Reported_Gender, yellow)

        if geno.value in problems:
            fill(clas, yellow)
            fill(geno, yellow)
        elif geno.value in problems_cnv:
            fill(clas, yellow)
            fill(geno, yellow)
            try:
                if int(bf.value) < 15:
                    fill(bf, yellow)
                else:
                    fill(Reads_Observed, yellow)
            except:
                pass
        elif geno.value in warnings:
            fill(clas, orange)
            fill(geno, orange)
        elif geno.value in georgian_druze:
            fill(clas, magneta)
            fill(geno, magneta)
        elif geno.value == non_reported_wt:
            fill(clas, green)
            fill(geno, green)
            clas.value = 'WT'
            continue

        for k,v in cfg.Classifictions.items():
            if geno.value in v:
                clas.value = k
                break

        if clas.value == 'HOM' or clas.value == 'SAMPLE FAILED':
            fill(geno, yellow)
            fill(clas, yellow)

        try:
            if int(gqx.value) < 15:
                fill(gqx, yellow)

            if float(altVar.value) < 30 or float(altVar.value) > 70:
                fill(altVar, yellow)
        except:
            pass

    for cells in ws.iter_rows(min_row=6, min_col=1, max_col=34):
        sample_cell = cells[0]
        clas = cells[9]
        geno = cells[12]
        if clas.value == 'No mutations identified (WT).':
            clas.value = 'WT'
        if clas.value in list(cfg.Classifictions_star.keys()):
            if geno.value not in cfg.Classifictions_star[clas.value]:
                clas.value = f'{clas.value}*'

    ws.auto_filter.ref = "A5:AH5"
    ws.sheet_view.zoomScale = 85
    wb.save(out_file)

    # ws.unmerge_cells('B2:D2')
    # ws.unmerge_cells('B3:D3')
    # ws.unmerge_cells('F2:G2')
    # ws.unmerge_cells('F3:G3')
    #
    # ws['B2'] = None
    # ws['E2'] = None
    # ws['F2'] = None
    # ws['B3'] = None
    # ws['E3'] = None
    # ws['F3'] = None
    #
    # r = ws.max_row
    # ws.move_range("C5:AH{}".format(r), cols=-1)
    # ws.move_range("C3:C3", cols=-1)
    # ws.move_range("D2:D2", cols=-1)
    # ws.move_range("F2:F2", rows=-1)
    # ws.move_range("C2:C2", cols=-1, rows=-1)
    # ws.move_range("A5:AH{}".format(r), rows=2)
    #
    # ws['B1'].value = 'Data Analysis Version'
    # ws['B3'].value = Analysis_Version
    # ws['D1'].value = 'Run Name'
    # ws['D3'].value = "{}".format(run)
    # ws['F1'].value = "Analysis Date"
    # ws['F3'].value = time.strftime("%d-%B-%Y")
    #
    # with open("{}.csv".format(out_file.split('.xlsx')[0]), 'w', newline='') as f:
    #     c = csv.writer(f)
    #     for r in ws.rows:
    #         c.writerow([cell.value for cell in r])

    return out_file, run


def AddIgvLink(dataframe, input_path, my_screen_version):
    if not input_path.endswith('\\'):
        short_input_path = input_path + '\\'
    else:
        short_input_path = input_path

    anno_file = pickle.load(open(cfg.anno_pkl, 'rb'))
    anno = pd.DataFrame(anno_file)
    dataframe["IGV Link (open IGV before)"] = ""
    for (i, row) in dataframe.iterrows():
        relevant_agid = anno[anno["AGID"] == row.loc['AGID']]
        locus = ""
        if relevant_agid.empty:
            if row.loc['Mutation'] != '---':
                match = re.findall(r"chr[0-9XY]{0,2}:g?\.?[0-9]+\-?[0-9]*", row.loc['Mutation'])
                if match != []:
                    locus = match[0].replace("g.", "")
        else:
            locus = str(relevant_agid["Chr"].iloc[0]) + ":" + str(int(relevant_agid["Position"].iloc[0]))
        if locus != "":
            sample_number = row.name
            for file in os.listdir(short_input_path):
                filename = str(file)
                if filename.startswith(str(sample_number)) and filename.endswith("bam"):
                    bamfile = filename
            link = 'http://localhost:60151/load?file={0}&locus={1}&genome={2}'.format((short_input_path + bamfile),
                                                                                      locus.strip(), "hg19")
            row.loc["IGV Link (open IGV before)"] = '=HYPERLINK(\"' + link + '\", \"' + str(locus.strip() + '\")')

    # in note - sub-script for automatically opening IGV application
    '''
    variables = os.environ['PATH'].split(';')
    potential_paths = [variable for variable in variables if re.match(".*IGV\_[0-9.]+\\\\$",variable)]
    flag = True
    try:
        igv_path = potential_paths[0]
    except:
        try:
            folder = [dir for dir in os.listdir("C:\\Program Files\\") if dir.startswith("IGV")]
            igv_path = ("C:\\Program Files\\" + folder[0] + "\\")
        except:
            flag = False
            print ("Couldn't find IGV installation path on the computer.")

     if flag:
        devnull = open(os.devnull, 'w')
        subprocess.run(igv_path+'igv.bat', stdout=devnull, stderr=devnull)  #only igv.bat is not working cause depends on current path
     print("added IGV links to sample summary")
     '''
    return dataframe


if __name__ == "__main__":
    original = ""
    target = ""
    # df = pd.DataFrame()

    # shutil.copyfile(original, target)
    paths = {'BAM_PATH': r'C:/Users/hadas/AGcloud/AGshared/Gamidor/Capture_Panel/HospitalRuns/Zer/201125_MN00742_0107_A000H3C7YJ', 'SAMPLE_DICT': {'S1': ['2305', 'Bedouin'], 'S2': ['2306', 'Bedouin'], 'S3': ['2314', 'Bedouin'], 'S4': ['2315', 'Bedouin'], 'S5': ['2336', 'Bedouin'], 'S6': ['2337', 'Bedouin'], 'S7': ['2338', 'Bedouin'], 'S8': ['2339', 'Bedouin'], 'S9': ['2340', 'Bedouin'], 'S10': ['2341', 'Bedouin'], 'S11': ['2347', 'Bedouin'], 'S12': ['2348', 'Bedouin'], 'S13': ['2349', 'Bedouin'], 'S14': ['2350', 'Bedouin'], 'S15': ['2351', 'Bedouin'], 'S16': ['2353', 'Bedouin'], 'S17': ['2354', 'Bedouin'], 'S18': ['2355', 'Bedouin'], 'S19': ['2357', 'Bedouin'], 'S20': ['2358', 'Bedouin'], 'S21': ['2359', 'Bedouin'], 'S22': ['2363', 'Bedouin'], 'S23': ['2364', 'Bedouin'], 'S24': ['2365', 'Bedouin'], 'S25': ['2367', 'Bedouin'], 'S26': ['2368', 'Bedouin'], 'S27': ['2369', 'Bedouin'], 'S28': ['2370', 'Bedouin'], 'S29': ['2371', 'Bedouin'], 'S30': ['2374', 'Bedouin'], 'S31': ['2375', 'Bedouin'], 'S32': ['2381', 'Bedouin'], 'S33': ['2382', 'Bedouin'], 'S34': ['2383', 'Bedouin'], 'S35': ['2389', 'Bedouin'], 'S36': ['2390', 'Bedouin'], 'S37': ['2398', 'Bedouin'], 'S38': ['2399', 'Bedouin'], 'S39': ['2400', 'Bedouin'], 'S40': ['2436', 'Bedouin']},
    'DIR_TREE': [r'C:/Users/hadas/AGcloud/AGshared/Gamidor/Capture_Panel/HospitalRuns/Zer/201125_MN00742_0107_A000H3C7YJ/MyScreen_Analysis_v2.1_RESULTS', r'C:/Users/hadas/AGcloud/AGshared/Gamidor/Capture_Panel/HospitalRuns/Zer/201125_MN00742_0107_A000H3C7YJ/MyScreen_Analysis_v2.1_RESULTS/Info', r'C:/Users/hadas/AGcloud/AGshared/Gamidor/Capture_Panel/HospitalRuns/Zer/201125_MN00742_0107_A000H3C7YJ/MyScreen_Analysis_v2.1_RESULTS/Info/Genotyping', r'C:/Users/hadas/AGcloud/AGshared/Gamidor/Capture_Panel/HospitalRuns/Zer/201125_MN00742_0107_A000H3C7YJ/MyScreen_Analysis_v2.1_RESULTS/Info/Genotyping/Logs', r'C:/Users/hadas/AGcloud/AGshared/Gamidor/Capture_Panel/HospitalRuns/Zer/201125_MN00742_0107_A000H3C7YJ/MyScreen_Analysis_v2.1_RESULTS/Info/Genotyping/Logs/PiscesLogs', r'C:/Users/hadas/AGcloud/AGshared/Gamidor/Capture_Panel/HospitalRuns/Zer/201125_MN00742_0107_A000H3C7YJ/MyScreen_Analysis_v2.1_RESULTS/Info/CNV', r'C:/Users/hadas/AGcloud/AGshared/Gamidor/Capture_Panel/HospitalRuns/Zer/201125_MN00742_0107_A000H3C7YJ/MyScreen_Analysis_v2.1_RESULTS/Info/CNV/Logs'],
    'EXTRA_INFO_PATH': False, 'RUN_NAME': '201013_MN00937_0047_A000H352JY', 'Hospital': 'Zer'}

    df = tools.decompress_pickle(r'c:\Users\hadas\AGcloud\AGshared\Gamidor\Capture_Panel\HospitalRuns\Zer\201125_MN00742_0107_A000H3C7YJ\MyScreen_Analysis_v2.1_RESULTS\Info\summary.pbz2')
    excel_formatter(df, paths, "MyScreen_Analysis_v2.1")
