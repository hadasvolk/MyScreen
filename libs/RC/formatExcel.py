import pandas as pd
import csv
import openpyxl
import openpyxl.styles as opstyl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, \
    colors  # for IGV link (only colors)
import matplotlib
import matplotlib.cm as cm
import numpy as np
import random
import string
from openpyxl import load_workbook
from itertools import islice
import glob
import shutil
import datetime
import re  # for igv link
import os  # for igv link
import subprocess  # for igv link
import pickle  # for igv link

import cfg

time = datetime.datetime.now()
date = datetime.datetime.now().strftime("%d-%m-%Y")

black = 'FF000000'
red = 'FFFF0028'
yellow = 'ffff00'
orange = 'ffa500'
cyan = '00FFFF'
grey = 'E0E0E0'
magneta = 'dce6f1'

thin = Side(border_style="thin", color="000000")
thick = Side(border_style="thick", color="000000")
double = Side(border_style="double", color="000000")

problems = ['WT', 'WT-Problem', 'Problem', 'CNV-Problem', 'NO_CALL', 'CARRIER-Georgian',
            'CARRIER-Georgian-Problem', 'CARRIER-Problem',
            'CNV-Problem Big Del Boundaries Different as Reported']
warnings = ['Big Del Boundaries Different as Reported',
            'CNV Big Del Boundaries Different as Reported']


def fill(cell, color, type='solid'):
    cell.fill = opstyl.PatternFill(start_color=color,
                                   end_color=color,
                                   fill_type=type)


def adjust_df(df):
    # df.sort_values(by=['S'], inplace=True)
    run = df.iloc[1]['Date of Run']
    df.drop(columns=['Date of Run'], inplace=True)
    # df.drop(columns=['S'], inplace=True)

    cols = list(df.columns.values)
    change = ['Sex', 'Sample Source', 'Mother Ethnicity', 'Father Ethnicity',
              'Partner Sample', 'AGID']
    for col in change:
        cols.pop(cols.index(col))
    df = df[cols + change]
    df.rename(columns={"Sex": "Analyzed Gender", "Gender": "Reported Gender",
                       "N.comp": "N comp", "Custom.first": "Custom First",
                       "Custom.last": "Custom Last", "Reads.expected": "Reads Expected",
                       "Reads.observed": "Reads Observed", "Reads.ratio": "Reads Ratio"}, inplace=True)
    df = df[['Sample', 'Disease', 'Gene', 'Mutation', 'MOH', 'Ethnicity', 'Classification',
            'Clalit Disease Makat', 'Clalit Mutation Makat', 'Genotype', 'GQX',
            'Alt Variant Freq', 'Read Depth', 'Alt Read Depth', 'Allelic Depths',
            'Correlation', 'N comp', 'Custom First', 'Custom Last', 'BF',
            'Reads Expected', 'Reads Observed', 'Reads Ratio', 'Analyzed Gender',
            'Reported Gender', 'Sample Source', 'Mother Ethnicity', 'Father Ethnicity',
            'Partner Sample', 'AGID', 'IGV Link (open IGV before)']]
    return df, run


def header(cell, to_bold):
    if to_bold:
        cell.font = opstyl.Font(bold=True)
    cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    fill(cell, magneta)


def excel_formatter(df, write_path, Analysis_Version):
    # wb = load_workbook(filename=write_path)
    # ws = wb.active
    # data = ws.values
    # cols = next(data)[0:]
    # data = list(data)
    # idx = [r[0] for r in data]
    # data = (islice(r, 0, None) for r in data)
    # df = pd.DataFrame(data, index=idx, columns=cols)
    df, run = adjust_df(df)

    new_order = list(df.columns)
    if 'IGV Link (open IGV before)' in new_order:  # move 'IGV Link' column to the end of the table
        new_order.remove('IGV Link (open IGV before)')
        new_order.append('IGV Link (open IGV before)')
    df = df.reindex(columns=new_order)

    out_file = "{}/sample_summary-{}.xlsx".format(write_path, date)
    with pd.ExcelWriter(out_file) as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)

    wb = openpyxl.load_workbook(out_file)
    ws = wb["Sheet1"]

    float_rows = ['K', 'L', 'M', 'T', 'W', 'P']
    for row in range(3, df.shape[0] + 2):
        for char in float_rows:
            try:
                val = round(float(ws["{}{}".format(char, row)].value), 4)
                ws["{}{}".format(char, row)].value = val
                ws["{}{}".format(char, row)].alignment = Alignment(horizontal='left')
            except:
                pass

    int_rows = ['H', 'I', 'U', 'V']
    for row in range(3, df.shape[0] + 2):
        for char in int_rows:
            try:
                val = int(ws["{}{}".format(char, row)].value)
                ws["{}{}".format(char, row)].value = val
                ws["{}{}".format(char, row)].alignment = Alignment(horizontal='left')
            except:
                pass

    ws.insert_rows(1, amount=4)
    ws.merge_cells('C2:D2')
    ws.merge_cells('C3:D3')
    ws.merge_cells('E2:F2')
    ws.merge_cells('E3:F3')

    ws['B2'] = "Data Analysis Version"
    ws['C2'] = "Run Name"
    ws['E2'] = "Analysis Date"
    ws['B3'] = Analysis_Version
    ws['C3'] = "{}".format(run)
    ws['E3'] = time.strftime("%d-%B-%Y")

    header_cells = ['B2', 'C2', 'D2', 'E2', 'F2']
    info_cells = ['B3', 'C3', 'D3', 'F3', 'E3']
    for cell in header_cells:
        header(ws[cell], True)
    for cell in info_cells:
        header(ws[cell], False)
    try:
        img = openpyxl.drawing.image.Image(cfg.MyScreen_Summary)
        img.anchor = 'H1'
        ws.add_image(img)
    except:
        ws['A2'] = "MyScreen"

    ws.freeze_panes = 'B6'
    letters = string.ascii_uppercase[7:26]
    for i in letters:
        ws.column_dimensions[i].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['M'].width = 8
    ws.column_dimensions['V'].width = 8
    ws.column_dimensions['X'].width = 10
    ws.column_dimensions['Y'].width = 10
    ws.column_dimensions['AB'].width = 10
    ws.column_dimensions['AE'].width = 25

    samples = []
    for col in ws['A']:
        samples.append(col.value)
    samples.pop(0)
    samples.pop(0)
    samples = list(dict.fromkeys(samples))[::2]

    for cells in ws.iter_rows(min_row=5, min_col=1, max_col=31):
        sample_cell = cells[0]
        clas = cells[6]
        geno = cells[9]
        gqx = cells[10]
        altVar = cells[11]
        mut = cells[3]
        ints = [cells[8], cells[7], cells[14], cells[15], cells[16], cells[17], cells[18], cells[19], cells[20], cells[21]]

        mut.alignment = Alignment(wrap_text=True)
        for i in range(3, 6):
            cells[i].alignment = Alignment(wrap_text=True)

        sample_cell.font = opstyl.Font(bold=True)
        sample_cell.border = Border(top=double, left=thin, right=thick, bottom=double)

        for each_cell in cells[1:]:
            try:
                each_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
            except Exception as e:
                print(e)
        if sample_cell.value in samples:
            for each_cell in cells:
                fill(each_cell, grey)

        if clas.value in problems:
            for each_cell in cells:
                fill(each_cell, yellow)
        elif clas.value in warnings:
            fill(clas, orange)

        if geno.value == 'hom' or geno.value == 'hom - With soft-clipped reads':
            fill(geno, orange)

        try:
            if int(gqx.value) < 15:
                for each_cell in cells:
                    fill(each_cell, yellow)

            if float(altVar.value) < 30 or float(altVar.value) > 70:
                fill(geno, yellow)
        except:
            pass

        for each_cell in ints:
            try:
                each_cell.value = each_cell.value.split('.')[0]
            except:
                pass

        if cells[30].row != 5:  # don't color the header
            cells[30].font = Font(u='single', color=colors.BLUE)  # add blue color to the link

    ws.sheet_view.zoomScale = 85
    wb.save(out_file)

    with open("{}.csv".format(out_file.split('.xlsx')[0]), 'w', newline='') as f:
        c = csv.writer(f)
        for r in ws.rows:
            c.writerow([cell.value for cell in r])


def AddIgvLink(dataframe, input_path, my_screen_version):
    if not input_path.endswith('\\'):
        short_input_path = input_path + '\\'
    else:
        short_input_path = input_path

    anno_file = pickle.load(open(cfg.anno_pkl, 'rb'))
    anno = pd.DataFrame(anno_file)
    dataframe["IGV Link (open IGV before)"] = ""
    for (i, row) in dataframe.iterrows():
        relevant_agid = anno[anno["AGID"] == row.loc['AGID']]
        locus = ""
        if relevant_agid.empty:
            if row.loc['Mutation'] != '---':
                match = re.findall(r"chr[0-9XY]{0,2}:g?\.?[0-9]+\-?[0-9]*", row.loc['Mutation'])
                if match != []:
                    locus = match[0].replace("g.", "")
        else:
            locus = str(relevant_agid["Chr"].iloc[0]) + ":" + str(relevant_agid["Position"].iloc[0])
        if locus != "":
            sample_number = row.Sample
            for file in os.listdir(short_input_path):
                filename = str(file)
                if filename.startswith(str(sample_number)) and filename.endswith("bam"):
                    bamfile = filename
            link = 'http://localhost:60151/load?file={0}&locus={1}&genome={2}'.format((short_input_path + bamfile),
                                                                                      locus.strip(), "hg19")
            row.loc["IGV Link (open IGV before)"] = '=HYPERLINK(\"' + link + '\", \"' + str(locus.strip() + '\")')

    # in note - sub-script for automatically opening IGV application
    '''
    variables = os.environ['PATH'].split(';')
    potential_paths = [variable for variable in variables if re.match(".*IGV\_[0-9.]+\\\\$",variable)]
    flag = True
    try:
        igv_path = potential_paths[0]
    except:
        try:
            folder = [dir for dir in os.listdir("C:\\Program Files\\") if dir.startswith("IGV")]
            igv_path = ("C:\\Program Files\\" + folder[0] + "\\")
        except:
            flag = False
            print ("Couldn't find IGV installation path on the computer.")

     if flag:
        devnull = open(os.devnull, 'w')
        subprocess.run(igv_path+'igv.bat', stdout=devnull, stderr=devnull)  #only igv.bat is not working cause depends on current path
     print("added IGV links to sample summary")
     '''
    return dataframe


if __name__ == "__main__":
    original = ""
    target = ""
    df = pd.DataFrame()

    shutil.copyfile(original, target)
    excel_formatter(df, target, "V2.0.b")
